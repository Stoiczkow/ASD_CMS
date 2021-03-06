import kronos
import datetime
import glob
import os
import re
from openpyxl import Workbook
from django.core.exceptions import ObjectDoesNotExist
from .models import Interruption, Machine, Realization, DBName, Order


@kronos.register('* * * * *')
def process_black_box():
    today = str(datetime.datetime.now().date()).replace('-', '')
    os.chdir("/mnt")
    file_name = ''
    db_name = DBName.objects.get(pk=1).name
    for i in glob.glob("*.txt"):
        if i.startswith(today):
            file_name = i

    with open(file_name) as file:
        all_lines = file.readlines()

        machines = Machine.objects.all()
        for col in range(0, len(machines)):
            current_interruption = ''
            for j in range(1, len(all_lines)):
                line = all_lines[j]
                line_without_date = line[20:]
                found = re.findall(r'([0-9]+)', line_without_date)
                if not int(found[col]):
                    if not current_interruption:
                        current_interruption = str(line[0:19])
                        current_interruption_dt = datetime.datetime.strptime(
                            current_interruption, '%d.%m.%Y %H:%M:%S')
                    try:
                        interruption = Interruption.objects.using(db_name).get(
                            start_date=current_interruption_dt,
                            machine=machines[col])
                    except ObjectDoesNotExist:
                        interruption = Interruption.objects.using(db_name).create(
                            start_date=current_interruption_dt,
                            machine=machines[col])
                else:
                    try:
                        interruption = Interruption.objects.using(db_name).get(
                            start_date=datetime.datetime.strptime(
                                current_interruption, '%d.%m.%Y %H:%M:%S'),
                            machine=machines[col])
                        interruption.stop_date = datetime.datetime.strptime(
                            str(line[0:19]), '%d.%m.%Y %H:%M:%S')
                        interruption.save(using=db_name)
                    except (ObjectDoesNotExist, ValueError):
                        pass

                    current_interruption = ''


@kronos.register('* * * * *')
def fill_interruptions():
    db_name = DBName.objects.get(pk=1).name
    interruptions = Interruption.objects.using(db_name).filter(realization=None)
    for interruption in interruptions:
        if interruption.stop_date:
            try:
                realization = Realization.objects.using(db_name).get(
                    order__machine=interruption.machine,
                    start_date__lte=interruption.stop_date,
                    stop_date__gte=interruption.start_date)
            except ObjectDoesNotExist:
                continue

            interruption.realization = realization
            interruption.save(using=db_name)
        else:
            try:
                realization = Realization.objects.using(db_name).get(
                    order__machine=interruption.machine,
                    is_active=True)
                interruption.realization = realization
                interruption.save(using=db_name)
            except ObjectDoesNotExist:
                continue


@kronos.register('* * * * *')
def write_excel():
    db_name = DBName.objects.get(pk=1).name
    if db_name == 'excel':
        wb = Workbook()
        orders = wb.active
        orders.title = 'Zlecenia'
        orders.append(['Maszyna', 'Początek', "Koniec", 'Planowane wykonanie'])
        for order in Order.objects.using(db_name).all():
            orders.append([order.machine.name, order.start_date,
                           order.stop_date, order.planned])

        realizations = wb.create_sheet(title="Realizacje")

        realizations.append(['Numer zlecenia', 'Początek', 'Koniec',
                             'Straty', 'Pracownik'])
        for realization in Realization.objects.using(db_name).all():
            realizations.append([realization.order.order_id, realization.start_date,
                                 realization.stop_date, realization.waste, realization.user.username])

        interruptions = wb.create_sheet(title="Przestoje")
        interruptions.append(['Początek', 'Koniec', 'Realizacja', 'Przyczyna 1',
                              'Przyczyna 2', 'Przyczyna_3'])
        for interruption in Interruption.objects.using(db_name).all():
            interruptions.append([interruption.start_date, interruption.stop_date,
                                 interruption.realization, interruption.cause_1,
                                 interruption.cause_2, interruption.cause_3])
        wb.save('data.xlsx')
